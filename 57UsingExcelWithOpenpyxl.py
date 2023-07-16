import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = tk.Tk()
root.title('Excel - OpenPyxl')
root.geometry('500x800')

# Create Workbook Instanc
wb = Workbook()

# Load Existing workbook
wb = load_workbook('pizza.xlsx')

# Create Active Worksheet
ws = wb.active

# Create Variables for Columns
column_a = ws['A']
column_b = ws['B']

def get_column_a():
    lis = ''
    for cell in column_a:
        lis = f'{lis + cell.value}\n'
    label_a.config(text=lis)

def get_column_b():
    lis = ''
    for cell in column_b:
        lis = f'{lis + cell.value}\n'
    label_b.config(text=lis)

btn_a = tk.Button(root, text="Get Column A", command=get_column_a)
btn_a.pack(pady=20)

btn_b = tk.Button(root, text="Get Column B", command=get_column_b)
btn_b.pack(pady=20)

label_a = tk.Label(root, text="")
label_a.pack(pady=20)

label_b = tk.Label(root, text="")
label_b.pack(pady=20)

# Add Samarth to A7
ws['A7'] = "Samarth"
ws['B7'] = "Cheese"

# Save New File
wb.save('pizz2.xlsx')

root.mainloop()