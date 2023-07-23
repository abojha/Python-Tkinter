import tkinter as tk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


root = tk.Tk()
root.title('Add Excel to List Box')
root.geometry('400x300')


# my_list = ["One", "Two", "Three", "Four"]

def select(event):
    my_label.config(text=my_listbox.get(tk.ANCHOR))

my_listbox = tk.Listbox(root, width=45)
my_listbox.pack(pady=20)


# Create a wb
wb = load_workbook('pizza.xlsx')

# Set active worksheet
ws = wb.active

# Grab a Column of data
col_a = ws["A"]
col_b = ws["B"]

for item in col_a:
    my_listbox.insert(tk.END, item.value)

my_label = tk.Label(root, text="", font=("Helvetica", 20))
my_label.pack(pady=20)

# Create List Box Binding
my_listbox.bind("<ButtonRelease-1>", select)

root.mainloop()